import os
import shutil
from configparser import ConfigParser
from datetime import datetime
from hashlib import md5

import openpyxl

config_file = 'config.ini'


def file_exist(file_path):
    return os.path.exists(file_path)


def copy_file(src, dst):
    shutil.copy2(src, dst)


def append_datetime(file_path):
    current_datetime = datetime.now().strftime('%Y%m%d_%H%M%S')
    directory, filename = os.path.split(file_path)
    base_name, extension = os.path.splitext(filename)

    new_filename = f'{base_name}_{current_datetime}{extension}'
    new_file_path = os.path.join(directory, new_filename)
    return new_file_path


def file_name(file_path):
    directory, filename = os.path.split(file_path)
    return filename


def is_excel_file(file_path):
    if not file_exist(file_path):
        return False
    try:
        wb = openpyxl.load_workbook(file_path)
        wb.close()
        return True
    except openpyxl.utils.exceptions.InvalidFileException as e:
        return False


def get_excel_sheets(file_path):
    if is_excel_file(file_path):
        wb = openpyxl.load_workbook(file_path)
        sheet_names = wb.sheetnames
        wb.close()
        return sheet_names
    return []


def get_excel_cell_total(excel_file, sheet_name):
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb[sheet_name]
    total = 0
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value:
                total += 1
    wb.close()
    return total


def get_configuration(section, variable):
    config = ConfigParser()
    try:
        config.read(config_file)
        value = config.get(section, variable)
        return value
    except Exception as e:
        return None


def md5_sign(s, encoding='utf-8'):
    return md5(s.encode(encoding)).hexdigest()


if __name__ == '__main__':
    test_file = './test.xlsx'
    print(get_excel_cell_total(test_file, 'Sheet1'))

    print(append_datetime(test_file))

    if is_excel_file(test_file):
        sheets = get_excel_sheets(test_file)
        print(sheets)
    else:
        print('not excel file')

    print(get_configuration('baidu', 'appid'))
    print(get_configuration('baidu', 'appkey'))
